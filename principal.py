#!/usr/bin/env python

from datetime import date
from conexion import *
import sys
from PyQt4 import QtCore, QtGui, uic
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
import smtplib 
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart





# Cargar archivo .ui
form_class = uic.loadUiType("ingresos.ui")[0]

class VentanaPrincipal(QtGui.QMainWindow, form_class):
	def __init__ (self,parent=None):
		QtGui.QMainWindow.__init__(self, parent)
		self.setupUi(self)
		
		
		#PROPIEDADES DE LA VENTANA
		self.setWindowTitle("Habilitacion | Sistema")
		self.setWindowIcon(QtGui.QIcon('/icons/icon.png'))
		
		
		#define posicion de la ventana
		self.move(QtGui.QApplication.desktop().screen().rect().center()- self.rect().center())
		#self.btn_mini.clicked.connect(lambda:self.showMinimized())
		
		#Elimina bordes
		#self.setWindowFlags(self.windowFlags() | QtCore.Qt.FramelessWindowHint)

		
		
		#cambiar de paginas
		self.btn_ingresar.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_ingresar)) #cambia de pagina -> ingresar
		self.btn_buscar.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_buscar)) #cambia de pagina -> buscar
		self.btn_listar.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_listar)) #cambia de pagina -> listar
		self.btn_actualizar.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_actualizar)) #cambia de pagina -> registrar pagos
		self.btn_modifica.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_modificar)) #cambia de pagina -> modificar
		self.btn_devoluciones.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_devoluciones))
		self.inicio.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_inicio))
		self.btn_afiliados.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_altaafiliados))
		self.btn_cargalote.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_lotes))
		self.btn_buscaafiliado.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_busquedaafiliado))
		self.btn_comisiones.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_comisiones))
		self.btn_config.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_config))
		
		self.btn_import.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_importar))
		
				
		#funcion de los botones
		self.txt_dni.lostFocus.connect(self.traeafiliado)
		self.btn_salir.clicked.connect(self.fsalir)
		self.btn_registrar.clicked.connect(self.registrardatos)
		self.btn_registrar.clicked.connect(self.limpiar)
		
		self.btn_modificar.clicked.connect(self.modificar)
		self.btn_modificar.clicked.connect(self.limpiar)
		
		#PAGINA ACTUALIZAR
		self.btn_actbuscar.clicked.connect(self.recuperainfoactualizar)
		self.btn_update.clicked.connect(self.actualizar)
		self.btn_update.clicked.connect(lambda:self.txt_actregistro.setEnabled(True))
		self.btn_update.clicked.connect(self.limpiar_registropago)
		self.btn_consulta_foja.clicked.connect(self.consultarorden)
		
				
		#SE ACTIVAN BOTONES CLICKEANDO EL RADIOBUTTON
		self.rb_carga_masiva.toggled.connect(lambda:self.btn_examinar_pagos.setEnabled(True))
		self.rb_carga_masiva.toggled.connect(lambda:self.btn_importar_pagos.setEnabled(True))
		self.rb_carga_individual.toggled.connect(lambda:self.btn_examinar_pagos.setEnabled(False))
		self.rb_carga_individual.toggled.connect(lambda:self.btn_importar_pagos.setEnabled(False))
		self.btn_importar_pagos.clicked.connect(self.importarpagos)
		self.btn_examinar_pagos.clicked.connect(self.elegirarchivo_pagos)
		self.rb_varias.toggled.connect(lambda:self.txt_valor_ajuste.setEnabled(True))
		self.rb_varias.toggled.connect(lambda:self.btn_ajustar.setEnabled(True))
		self.rb_unica.toggled.connect(lambda:self.txt_valor_ajuste.setEnabled(False))
		self.rb_unica.toggled.connect(lambda:self.btn_ajustar.setEnabled(False))
		
		
	
		#funcion si txtactregistro pierde el foco
		#self.txt_actregistro.focusOutEvent(self.recuperainfoactualizar)
			
		
		#PAGINA MODIFICAR
		self.btn_liberadni.clicked.connect(lambda:self.txt_modificar_dni.setEnabled(True)) #libera campo de texto para luego actualizar
		self.btn_liberanombre.clicked.connect(lambda:self.txt_modificar_nombre.setEnabled(True))
		self.btn_liberaimporte.clicked.connect(lambda:self.txt_modificar_importe.setEnabled(True))
		self.btn_liberaconcepto.clicked.connect(lambda:self.txt_modificar_concepto.setEnabled(True))
		self.btn_liberafechapago.clicked.connect(lambda:self.txt_modificar_fechapago.setEnabled(True))
		self.btn_liberafechaprestacion.clicked.connect(lambda:self.txt_modificar_fechaprestacion.setEnabled(True))
		self.btn_liberanumtransferencia.clicked.connect(lambda:self.txt_modificar_transferencia.setEnabled(True))
		self.btn_liberaobs.clicked.connect(lambda:self.txt_modificar_obs.setEnabled(True))
		self.btn_liberacuenta.clicked.connect(lambda:self.txt_modificar_cuenta.setEnabled(True))
		self.btn_liberanuevoreg.clicked.connect(lambda:self.txt_modificar_nuevoreg.setEnabled(True))
		
		
		self.btn_busqueda.clicked.connect(self.busqueda)
		self.btn_eliminar_registros.clicked.connect(self.eliminarregistro)
		self.btn_busqueda.clicked.connect(self.limpiar)
		
		#self.btn_listarlistar.clicked.connect(self.listar)
		
		self.btn_exportar_busqueda.clicked.connect(self.exportar_busqueda)
		
		#PAGINA MODIFICAR
			
		self.btn_recuperar.clicked.connect(self.recuperainfomodificar)
		
		
		#PAGINA DEVOLUCIONES
		self.btn_verhistorial.clicked.connect(self.verhistorial)
		self.btn_devolver.clicked.connect(self.nuevadevolucion)
		self.btn_reingresar.clicked.connect(self.reingresarregistro)
		self.btn_observar.clicked.connect(self.nuevaobservacion)
		self.btn_reingresar.clicked.connect(self.limpiar)
		self.btn_devolver.clicked.connect(self.limpiar)	
		
		#PAGINA LISTAR
		
		self.btn_listarlistar.clicked.connect(self.listadotodos)
		self.btn_exportar.clicked.connect(self.exportarlistar)
		self.rb_portransferencias.toggled.connect(self.bloquearrblistar)
		self.rb_ingresos.toggled.connect(self.mostrarrblistar)
		self.btn_relacion.clicked.connect(self.relaciondegastos)
		self.btn_limpiarlistar.clicked.connect(self.limpiarlista)
		
		
		#pagina lotes
		self.btn_listarloteables.clicked.connect(self.listadoloteables)
		self.btn_lotear.clicked.connect(self.lotear)
		self.btn_exportar_lote.clicked.connect(self.exportarlote)
		self.btn_exportar_lote.clicked.connect(self.exportardetallelote)
		self.tb_lotes.doubleClicked.connect(self.eliminarloteado)
		
		
		#PAGINA ALTA AFILIADOS
		self.btn_registraralta.clicked.connect(self.altaafiliado)
		self.btn_registraralta.clicked.connect(self.limpiar)
		self.btn_eliminar_afiliado.clicked.connect(self.eliminarafiliados)
		
		#PAGINA BUSCAR AFILIADO
		self.btn_buscarafiliado.clicked.connect(self.buscarafiliado)
		
		#liberar campos de texto para luego actualizar
		self.btn_libera_ayn_afiliado.clicked.connect(lambda:self.txt_nombre_busafi.setEnabled(True))
		self.btn_liberacbu_afiliado.clicked.connect(lambda:self.txt_cbu_busafi.setEnabled(True))
		self.btn_liberatel_afiliado.clicked.connect(lambda:self.txt_telbusafi.setEnabled(True))
		self.btn_liberaemail_afiliado.clicked.connect(lambda:self.txt_email_busafi.setEnabled(True))
		self.btn_liberaalias_afiliado.clicked.connect(lambda:self.txt_aliasbusafi.setEnabled(True))
		self.btn_liberacuenta_afliado.clicked.connect(lambda:self.txt_cuenta_busafi.setEnabled(True))
		self.btn_liberacuil_afiliado.clicked.connect(lambda:self.txt_cuil_busafi.setEnabled(True))
		self.btn_liberaestado_afiliado.clicked.connect(lambda:self.combo_estado_busafi.setEnabled(True))
		
		self.btn_actualizar_afiliado.clicked.connect(self.updateafiliado)
		self.btn_actualizar_afiliado.clicked.connect(self.limpiar)

		
		#PAGINA COMISIONES
		self.btn_cargar_comision.clicked.connect(self.nuevacomision)
		self.btn_seleccionar_agente	.clicked.connect(self.vercomisiones)
		self.btn_eliminar_comision.clicked.connect(self.eliminar_comision)
		
		self.cbx_eneldia.toggled.connect(lambda:self.date_comision_regreso.setEnabled(True))
		self.btn_ajustar.clicked.connect(self.ajustar_comision)
		
		#propiedades de elementos_
		#Definir contenido ajustable en las tablas:
		#tabla actualizar
		headertb_actualizar = self.tb_actualizar.horizontalHeader()
		headertb_actualizar.setResizeMode(QtGui.QHeaderView.ResizeToContents)
		#tabla busqueda
		headertb_busqueda = self.tb_busqueda.horizontalHeader()
		headertb_busqueda.setResizeMode(QtGui.QHeaderView.ResizeToContents)
		#tabla devoluciones
		headertb_devoluciones = self.tb_devoluciones.horizontalHeader()
		headertb_devoluciones.setResizeMode(QtGui.QHeaderView.ResizeToContents)
		#tabla listar
		headertb_listar = self.tb_listar.horizontalHeader()
		headertb_listar.setResizeMode(QtGui.QHeaderView.ResizeToContents)
		#tabla lotes
		headertb_lotes = self.tb_lotes.horizontalHeader()
		headertb_lotes.setResizeMode(QtGui.QHeaderView.ResizeToContents)
		#tabla comisiones
		headertb_comisiones = self.tb_comisiones.horizontalHeader()
		headertb_comisiones.setResizeMode(QtGui.QHeaderView.ResizeToContents)
		
		
		
		
		self.fecha_hastalistar.setDate(date.today())
		self.fecha_desdelistar.setDate(date.today())
		self.rb_todos.setChecked(True)
		
		self.desde_lote.setDate(date.today())
		self.hasta_lote.setDate(date.today())
		self.txt_actfecha.setDate(date.today())
		self.date_comision.setDate(date.today())
		#self.date_comision.dateChanged(self.date_comision_regreso=self.date_comision)
		self.date_comision_regreso.setDate(date.today())
		
		#PAGINA CONFIG
		self.btn_agregar_agente.clicked.connect(self.nuevoagente)
		self.fecha_valor_comision.setDate(date.today())
		self.btn_definir_comision.clicked.connect(self.definirvalorcomision)
		
		
		#PAGINA IMPORTAR
		self.btn_importar_cuentas.clicked.connect(self.importarcuentas)
		self.btn_archivo_cuentas.clicked.connect(self.elegirarchivo)
		self.btn_examinar_registros.clicked.connect(self.elegirarchivo_registros)
		self.btn_importar_registros.clicked.connect(self.importar)
		
		
		
	def llenarcombos(self):
		
		agentes=q.traeragentes()
		
		
		k=0
		for i in agentes:
			
			self.cb_agente.addItem("".join(agentes[k]))
			k=k+1
		
		
	def traeafiliado(self):
		
		dni= int(self.txt_dni.text())
		nombre_afiliado="".join(q.getnombre_afiliado(dni))
		
		self.txt_ayn.setText(nombre_afiliado)
		estado = q.getEstado(dni)
		
		if estado[0]=='FALLECIDO':
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(3)
			msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
			msgBox.setWindowTitle("ATENCION")
			msgBox.setText("AFILIADO FALLECIDO")
			msgBox.exec_()
		
				
		
			
		
	def registrardatos(self):
		
		if self.txt_registro.text():
			
			registro=str(self.txt_registro.text())
			dni= int(self.txt_dni.text())
			fecha=str(date.today())		
			ayn=str(self.txt_ayn.text())
			importe=float(self.txt_importe.text())
			fechap= str(self.date_ingresos.text())
			concepto= str(self.cb_ingresos.currentText())
			estado="PENDIENTE"
			observaciones= str(self.txt_observaciones.text())
			if concepto == 'Ortopedia Blanda' or concepto == 'Subsidio de Sepelio':
				cuenta=str(30)
			else:
				cuenta=str(26)
		
						
			self.txt_registro.setFocus()
						
					
		
			if self.cbx_loteable.isChecked():
				duplicados=q.validarduplicados(registro)
				if duplicados == None:
				
					cuentaOK=q.validacuenta(dni) #VALIDA QUE TENGA UNA CUENTA, SI NO TIENE NO SE PUEDE LOTEAR
					
				
					if cuentaOK != None:
						q.insertarenlotes(registro,dni,importe,fecha)
						q.insertarenbd(registro,dni,importe,concepto,fecha,fechap,estado,observaciones,ayn,cuenta)
					
						#mensaje
						msgBox=QtGui.QMessageBox(self.centralwidget)
						msgBox.setIcon(1)
						msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
						msgBox.setWindowTitle("REGISTRO INGRESADO")
						msgBox.setText("EL INGRESO LOTEABLE HA SIDO EXITOSO")
						msgBox.exec_()
					else:
						#mensaje
						q.insertarenbd(registro,dni,importe,concepto,fecha,fechap,estado,observaciones,ayn,cuenta)
						msgBox=QtGui.QMessageBox(self.centralwidget)
						msgBox.setIcon(2)
						msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
						msgBox.setWindowTitle("REGISTRO NO LOTEABLE")
						msgBox.setText("EL AFILIADO NO POSEE CUENTA REGISTRADA")
						msgBox.exec_()
				else:
					#mensaje
					msgBox=QtGui.QMessageBox(self.centralwidget)
					msgBox.setIcon(3)
					msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
					msgBox.setWindowTitle(" * *ATENCION * * ")
					msgBox.setText("EL EXPEDIENTE YA ESTABA REGISTRADO.NUM-->: "+str(duplicados[0]))
					msgBox.exec_();
					
					
			else:
				duplicados=q.validarduplicados(registro)
			
				if duplicados == None:
					q.insertarenbd(registro,dni,importe,concepto,fecha,fechap,estado,observaciones,ayn,cuenta)
						
				
					#mensaje
					msgBox=QtGui.QMessageBox(self.centralwidget)
					msgBox.setIcon(1)
					msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
					msgBox.setWindowTitle("REGISTRO INGRESADO")
					msgBox.setText("EL INGRESO HA SIDO EXITOSO")
					msgBox.exec_()
				else:
				
				
					#mensaje
					msgBox=QtGui.QMessageBox(self.centralwidget)
					msgBox.setIcon(3)
					msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
					msgBox.setWindowTitle(" * *ATENCION * * ")
					msgBox.setText("EL EXPEDIENTE YA ESTABA REGISTRADO.NUM-->: "+str(duplicados[0]))
					msgBox.exec_();
		else:
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(3)
			msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
			msgBox.setWindowTitle("ERROR")
			msgBox.setText("CAMPO VACIO")
			msgBox.exec_()
		
		
				
			
			
			
	def busqueda(self):
		
		if self.txt_busqueda.text():
			
		
			if self.rb_pordni.isChecked():
				
				dnix = self.txt_busqueda.text()
		
				
				tablarecuperada=q.pordni(int(dnix))
				totalfilas=len(tablarecuperada)
				self.tb_busqueda.setRowCount(totalfilas) #cantidad de filas a llenar *debe ir si o si*
				fila=0
		
		
			
			
			elif self.rb_porregistro.isChecked():
				dnix = self.txt_busqueda.text()
				
				tablarecuperada=q.porregistro(str(dnix))
				totalfilas=len(tablarecuperada)
				self.tb_busqueda.setRowCount(totalfilas) #cantidad de filas a llenar *debe ir si o si*
				fila=0
			
			elif self.rb_pornombre.isChecked():
				dnix = self.txt_busqueda.text()
				
				tablarecuperada=q.pornombre(str(dnix))
				totalfilas=len(tablarecuperada)
				self.tb_busqueda.setRowCount(totalfilas) #cantidad de filas a llenar *debe ir si o si*
				fila=0
			
			
		
			
			
			for i in tablarecuperada:
				self.tb_busqueda.setItem(fila,0,QtGui.QTableWidgetItem(str(i[4])))#GUI INGRESO
				self.tb_busqueda.setItem(fila,1,QtGui.QTableWidgetItem(str(i[0])))#GUI REGISTRO
				self.tb_busqueda.setItem(fila,2,QtGui.QTableWidgetItem(str(i[1])))#GUI DNI
				self.tb_busqueda.setItem(fila,3,QtGui.QTableWidgetItem(str(i[2])))#GUI IMPORTE
				
				if i[7] =="TRANSFERIDO":
					
					
					self.tb_busqueda.setItem(fila, 4, QtGui.QTableWidgetItem("TRANSFERIDO"))
					self.tb_busqueda.item(fila,4).setBackground(QtGui.QColor(112, 255, 110))
							
					
				elif i[7] =="PENDIENTE":
					self.tb_busqueda.setItem(fila, 4, QtGui.QTableWidgetItem("PENDIENTE"))
					self.tb_busqueda.item(fila,4).setBackground(QtGui.QColor(255, 143, 145))
					
				elif i[7] =="OBSERVADO":
					self.tb_busqueda.setItem(fila, 4, QtGui.QTableWidgetItem("OBSERVADO"))
					self.tb_busqueda.item(fila,4).setBackground(QtGui.QColor(255, 255, 127))
				elif i[7] =="DEVUELTO":
					self.tb_busqueda.setItem(fila, 4, QtGui.QTableWidgetItem("DEVUELTO"))
					self.tb_busqueda.item(fila,4).setBackground(QtGui.QColor(255, 145, 55))
					
					
				
				self.tb_busqueda.setItem(fila,5,QtGui.QTableWidgetItem(str(i[10])))#GUI NOMBRE
				self.tb_busqueda.setItem(fila,6,QtGui.QTableWidgetItem(str(i[8])))#GUI N TRANSFERENCIA
				self.tb_busqueda.setItem(fila,7,QtGui.QTableWidgetItem(str(i[5])))#GUI FECHA PAGO
				self.tb_busqueda.setItem(fila,8,QtGui.QTableWidgetItem(str(i[9])))#GUI OBS
				self.tb_busqueda.setItem(fila,9,QtGui.QTableWidgetItem(str(i[3])))#GUI CONCEPTO
				self.tb_busqueda.setItem(fila,10,QtGui.QTableWidgetItem(str(i[6])))#GUI FECHA PRESTACION
				self.tb_busqueda.setItem(fila,11,QtGui.QTableWidgetItem(str(i[13])))#GUI ORDEN
				
				
				
				fila+=1
			
			
			
			
		else:
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(3)
			msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
			msgBox.setWindowTitle("ERROR")
			msgBox.setText("CAMPO VACIO")
			msgBox.exec_()
		
		
	def exportar_busqueda(self):
		lista=[]
		cantidad_filas=self.tb_busqueda.rowCount()
		
		
		book = Workbook()
		sheet = book.active
		
		sheet['A1']="REPORTE"
		sheet['B1']="DNI"
		sheet['C1']="IMPORTE"
		sheet['D1']="ESTADO"
		sheet['E1']="NOMBRE"
		sheet['F1']="COMPROBANTE"
		sheet['G1']="FECHA PAGO"
		sheet['H1']="OBS"
		sheet['I1']="RUBRO"
		sheet['J1']="MES PRESTACION"
		
		
		
		for row in range (cantidad_filas):
			linea=[]
			for column in range (1,11):
				
				items= self.tb_busqueda.item(row, column)
				
				linea.append(str(items.text()))
				
				
			lista.append(linea)	
			
			
				
		
		book = Workbook()
		sheet = book.active
		
		
		
		
		
		for i in lista:
			
			sheet.append(i)
			
		
		
		book.save('/home/usuario/busqueda_exportada.xlsx')
		
		msgBox=QtGui.QMessageBox(self.centralwidget)
		msgBox.setIcon(1)
		msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
		msgBox.setWindowTitle("AVISO")
		msgBox.setText("ARCHIVO EXPORTADO")
		msgBox.exec_()
				
			
			
		
		
	def actualizar(self):
		
		if self.txt_actregistro.text():
			
			regx = str(self.txt_actregistro.text())
			fechapago= str(self.txt_actfecha.text())
			mes=str(fechapago[5:7])
			anio=str(fechapago[0:4])
		
			cuenta=str(self.txt_cuenta.text())
			if self.cbx_fojamanual.isChecked():
				orden=str(self.txt_actorden.text())
			else:			
				orden=q.recuperaultimoorden(mes,cuenta,anio)
		
			transferencia= str(self.txt_acttransferencia.text())
			estado = "TRANSFERIDO"
			indice = int(self.txt_actindice.text())
		
			q.actualizapagoenbd(regx,fechapago,transferencia,estado,indice,int("".join(map(str,orden))),mes,anio)
			self.noti_actualizar.setText("Pago Registrado bajo la FOJA: "+str(int("".join(map(str,orden)))))
			self.txt_actregistro.setFocus()
			self.tb_actualizar.clear()
			self.consultarorden()
		else:
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(3)
			msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
			msgBox.setWindowTitle("ERROR")
			msgBox.setText("CAMPO VACIO")
			msgBox.exec_()
		
		
		
		
	def recuperainfoactualizar(self):
		
		if self.txt_actregistro.text():
			#self.avisomail()
			self.noti_actualizar.setText("")
			
			registro = str(self.txt_actregistro.text())
			listarecuperada= q.recuperadatosenbd(registro)
			totalfilas=len(listarecuperada)
			self.tb_actualizar.setRowCount(totalfilas) #cantidad de filas a llenar *debe ir si o si*
			
			
			fila =0
			for i in listarecuperada:
				if totalfilas ==1:
					self.txt_actindice.setText(str(i[5]))			
				
				self.tb_actualizar.setItem(fila,0,QtGui.QTableWidgetItem(str(i[0])))
				self.tb_actualizar.setItem(fila,1,QtGui.QTableWidgetItem(str(i[1])))
				self.tb_actualizar.setItem(fila,2,QtGui.QTableWidgetItem(str(i[2])))
				self.tb_actualizar.setItem(fila,3,QtGui.QTableWidgetItem(str(i[3])))
				self.tb_actualizar.setItem(fila,4,QtGui.QTableWidgetItem(str(i[4])))
				self.tb_actualizar.setItem(fila,5,QtGui.QTableWidgetItem(str(i[5])))
				
				self.txt_cuenta.setText(str(i[6]))
			
				
				
				fila= fila +1
				
				
		else:
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(3)
			msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
			msgBox.setWindowTitle("ERROR")
			msgBox.setText("CAMPO VACIO")
			msgBox.exec_()
			
	def consultarorden(self):
		
		fecha=str(date.today())
		mes=str(fecha[5:7])
		anio=str(fecha[0:4])
		
		
		orden26=q.recuperaultimoorden(str(mes),str(26),anio)
		orden30=q.recuperaultimoorden(str(mes),str(30),anio)
		
		
		self.signal_foja26.setText("".join(map(str,orden26)))
		self.signal_foja30.setText("".join(map(str,orden30)))	
		
	def recuperainfomodificar(self):
		
		if self.txt_modificar_registro.text():
			
			
			registro= str(self.txt_modificar_registro.text())
			
			
			listarecuperada = q.recuperatodoenbd(registro)
			for i in listarecuperada:
				self.txt_modificar_dni.setText(str(i[0]))
				self.txt_modificar_nombre.setText(str(i[1]))
				self.txt_modificar_importe.setText(str(i[2]))
				self.txt_modificar_concepto.setText(str(i[3]))
				self.txt_modificar_fechapago.setText(str(i[4]))
				self.txt_modificar_fechaprestacion.setText(str(i[5]))
				self.txt_modificar_transferencia.setText(str(i[6]))
				self.txt_modificar_obs.setText(str(i[7]))
				self.txt_modificar_cuenta.setText(str(i[8]))
				self.txt_modificar_nuevoreg.setText(str(i[9]))
				self.signal_indice.setText(str(i[10]))
				
				
		else:
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(3)
			msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
			msgBox.setWindowTitle("ERROR")
			msgBox.setText("CAMPO VACIO")
			msgBox.exec_()		
				
	def llenarcamposmodificar(self):
		fila=self.tb_modificar.currentRow()
		item=self.tb_modificar.item(fila,5)
		print str(item.text())
		indice=int(item.text())
		
		listarecuperada=q.recuperatodoconindice(indice)
		#HACER SIN UN FOR
		for i in listarecuperada:
				self.txt_modificar_dni.setText(str(i[0]))
				self.txt_modificar_nombre.setText(str(i[1]))
				self.txt_modificar_importe.setText(str(i[2]))
				self.txt_modificar_concepto.setText(str(i[3]))
				self.txt_modificar_fechapago.setText(str(i[4]))
				self.txt_modificar_fechaprestacion.setText(str(i[5]))
				self.txt_modificar_transferencia.setText(str(i[6]))
				self.txt_modificar_obs.setText(str(i[7]))
		
		
			
	def modificar(self):
		
		dni = int(self.txt_modificar_dni.text())
		nombre = str(self.txt_modificar_nombre.text())
		importe = float(self.txt_modificar_importe.text())
		concepto = str(self.txt_modificar_concepto.text())
		fechapago = str(self.txt_modificar_fechapago.text())
		fechaprestacion = str(self.txt_modificar_fechaprestacion.text())
		transferencia = self.txt_modificar_transferencia.text()
		obs = str(self.txt_modificar_obs.text())
		nuevoreg=str(self.txt_modificar_nuevoreg.text())
		cuenta=str(self.txt_modificar_cuenta.text())
		indice=int(self.signal_indice.text())
		
		#validar vacios
		
		if dni =="":
			dni=0
		if nombre=="":
			nombre="-"
		if importe=="":
			importe=0
		if concepto=="":
			concepto="-"
		if fechapago=="":
			fechapago=""
		if fechaprestacion =="":
			fechaprestacion=""
		if transferencia =="" or transferencia =="None":
			transferencia=0
		if obs=="":
			obs=""
			
			
				
			
		q.actualizatodoenbd(dni,nombre,importe,concepto,str(fechapago),fechaprestacion,int(transferencia),obs,cuenta,nuevoreg,indice)
		if self.cbx_modificaenlote.isChecked():
			registro=str(self.txt_modificar_registro.text())
			q.actualizarenlotes(registro,importe)
		
		#MENSAJE OK
		msgBox=QtGui.QMessageBox(self.centralwidget)
		msgBox.setIcon(1)
		msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
		msgBox.setWindowTitle("MODIFICACION")
		msgBox.setText("REGISTRO MODIFICADO")
		msgBox.exec_()	
		#LIMPIA FORMULARIO
		self.txt_modificar_registro.setText("")
		self.txt_modificar_dni.setText("")
		self.txt_modificar_nombre.setText("")
		self.txt_modificar_importe.setText("")
		self.txt_modificar_concepto.setText("")
		self.txt_modificar_fechapago.setText("")
		self.txt_modificar_fechaprestacion.setText("")
		self.txt_modificar_transferencia.setText("")
		self.txt_modificar_obs.setText("")
		
		
	
	def eliminarregistro(self):
		msgBox=QtGui.QMessageBox(self.centralwidget)
		msgBox.setIcon(2)
		msgBox.setStandardButtons(QtGui.QMessageBox.Yes | QtGui.QMessageBox.Cancel | QtGui.QMessageBox.No)
		
		msgBox.setWindowTitle(" * * ATENCION * * ")
		msgBox.setText("DESEA ELIMINAR EL REGISTRO?")
		r= msgBox.exec_()
		
		
		
		if r==16384:
			
					
			registro= str(self.txt_modificar_registro.text())
			dni=int(self.txt_modificar_dni.text())
			q.eliminarregistros(registro,dni)
			
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setWindowTitle(" * * ATENCION * * ")
			msgBox.setText("REGISTRO ELIMINADO")
			resp= msgBox.exec_()
			
		elif r==4194304:
			pass
		elif r==65536:
			pass
			
	
		self.txt_modificar_registro.setText("")
		self.txt_modificar_dni.setText("")
		self.txt_modificar_nombre.setText("")
		self.txt_modificar_importe.setText("")
		self.txt_modificar_concepto.setText("")
		self.txt_modificar_fechapago.setText("")
		self.txt_modificar_fechaprestacion.setText("")
		self.txt_modificar_transferencia.setText("")
		self.txt_modificar_obs.setText("")	
		self.txt_modificar_cuenta.setText("")
		self.self.signal_indice.setText("")
			
		
	def verhistorial(self):
		
		if self.txt_devregistro.text():
			
		
			registro=str(self.txt_devregistro.text())
			
			tablarecuperada=q.recuperahistorial(registro)
			if tablarecuperada:
				totalfilas= len(tablarecuperada)
				self.tb_devoluciones.setRowCount(totalfilas) #cantidad de filas a llenar *debe ir si o si*
				fila =0
				
				for i in tablarecuperada:
					self.tb_devoluciones.setItem(fila,0,QtGui.QTableWidgetItem(str(i[0])))#REG
					self.tb_devoluciones.setItem(fila,1,QtGui.QTableWidgetItem(str(i[1])))#DEVOLUCION
					self.tb_devoluciones.setItem(fila,2,QtGui.QTableWidgetItem(str(i[2])))#REINGRESO
					self.tb_devoluciones.setItem(fila,3,QtGui.QTableWidgetItem(str(i[3])))#MOTIVO
					self.tb_devoluciones.setItem(fila,4,QtGui.QTableWidgetItem(str(i[4])))#DESTINO
					fila = fila+1
			else:
				msgBox=QtGui.QMessageBox(self.centralwidget)
				msgBox.setIcon(1)
				msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
				msgBox.setWindowTitle("SIN HISTORIAL")
				msgBox.setText("EL EXPEDIENTE INGRESADO NO POSEE HISTORIAL REGISTRADO")
				msgBox.exec_()
				
		else:
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(3)
			msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
			msgBox.setWindowTitle("ERROR")
			msgBox.setText("CAMPO VACIO")
			msgBox.exec_()
			
	def nuevadevolucion(self):
		
		if self.txt_devregistro.text():
			registro=str(self.txt_devregistro.text())
			motivo=str(self.txt_dev_motivo.text())
			destino=str(self.txt_dev_destino.currentText())
			fechadevolucion=str(date.today())
			
			validado=q.validaregistro(registro)
			
			if validado != None:
			
				q.registrardevolucion(registro,fechadevolucion,motivo,destino)
				q.observa_estado(registro,'DEVUELTO')
				#pagina devoluciones
				self.txt_devregistro.setText("")
				self.txt_dev_motivo.setText("")
				
				msgBox=QtGui.QMessageBox(self.centralwidget)
				msgBox.setIcon(2)
				msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
				msgBox.setWindowTitle("DEVOLUCION")
				msgBox.setText("EXPEDIENTE DEVUELTO CORRECTAMENTE")
				msgBox.exec_()
			else:
				msgBox=QtGui.QMessageBox(self.centralwidget)
				msgBox.setIcon(3)
				msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
				msgBox.setWindowTitle("ERROR")
				msgBox.setText("NO EXISTE ESE NUMERO DE REGISTRO. CORREGIR")
				msgBox.exec_()
				
				
		else:
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(3)
			msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
			msgBox.setWindowTitle("ERROR")
			msgBox.setText("CAMPO VACIO")
			msgBox.exec_()
	
	def nuevaobservacion(self):
		
		if self.txt_devregistro.text():
			registro=str(self.txt_devregistro.text())
			motivo=str(self.txt_dev_motivo.text())
			destino='HABILITACION'
			fechadevolucion=str(date.today())
			if self.txt_dev_motivo.text():
				
				validado=q.validaregistro(registro)
				if validado != None:
					q.registrardevolucion(registro,fechadevolucion,motivo,destino)
					q.observa_estado(registro,'OBSERVADO')
					#pagina devoluciones
					self.txt_devregistro.setText("")
					self.txt_dev_motivo.setText("")
					
					msgBox=QtGui.QMessageBox(self.centralwidget)
					msgBox.setIcon(2)
					msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
					msgBox.setWindowTitle("MENSAJE")
					msgBox.setText("EXPEDIENTE OBSERVADO CORRECTAMENTE")
					msgBox.exec_()
				else:
					msgBox=QtGui.QMessageBox(self.centralwidget)
					msgBox.setIcon(3)
					msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
					msgBox.setWindowTitle("ERROR")
					msgBox.setText("NO EXISTE ESE NUMERO DE REGISTRO. CORREGIR")
					msgBox.exec_()
					
					
			else:
				msgBox=QtGui.QMessageBox(self.centralwidget)
				msgBox.setIcon(3)
				msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
				msgBox.setWindowTitle("SIN MOTIVO")
				msgBox.setText("INGRESE MOTIVO")
				msgBox.exec_()
				
		else:
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(3)
			msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
			msgBox.setWindowTitle("ERROR")
			msgBox.setText("CAMPO VACIO")
			msgBox.exec_()
		
		
	def reingresarregistro(self):
		registro=str(self.txt_devregistro.text())
		fechareingreso=str(date.today())
		destino = "HABILITACION"
		
		validado=q.validaregistro(registro)
		if validado != None:
			q.reingresar(registro,fechareingreso,destino)
			q.observa_estado(registro,'PENDIENTE')
			
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(1)
			msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
			msgBox.setWindowTitle("REINGRESO")
			msgBox.setText("EL EXPEDIENTE FUE REINGRESADO A LA OFICINA CORRECTAMENTE CON ESTADO PENDIENTE")
			msgBox.exec_()
		else:
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(3)
			msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
			msgBox.setWindowTitle("ERROR")
			msgBox.setText("NO EXISTE ESE NUMERO DE REGISTRO. CORREGIR")
			msgBox.exec_()
		
		
	def listadotodos(self):
		
		if self.rb_26.isChecked():
			cuenta='26'
		elif self.rb_30.isChecked():
			cuenta='30'
		else:
			cuenta='%'
		
		
		concepto= str(self.cb_concepto_listar.currentText())
		if concepto == '':
			concepto='%'
		
		if self.rb_todos.isChecked():
			estado='%'
			fechaini= str(self.fecha_desdelistar.text())
			fechafin=str(self.fecha_hastalistar.text())
			
			if self.rb_ingresos.isChecked():
				tablarecuperada= q.listarporfechaingreso(fechaini,fechafin,concepto,cuenta,estado)
			else:
				tablarecuperada= q.listarporfechatransferencias(fechaini,fechafin,concepto,cuenta)
		
			totalfilas= len(tablarecuperada)
			self.tb_listar.setRowCount(totalfilas) #cantidad de filas a llenar *debe ir si o si*
			fila =0
		elif self.rb_transferidos.isChecked():
			estado='TRANSFERIDO'
			fechaini= str(self.fecha_desdelistar.text())
			fechafin=str(self.fecha_hastalistar.text())
			
			tablarecuperada= q.listarporfechaingreso(fechaini,fechafin,concepto,cuenta,estado)
		
			totalfilas= len(tablarecuperada)
			self.tb_listar.setRowCount(totalfilas) #cantidad de filas a llenar *debe ir si o si*
			fila =0
		elif self.rb_pendientes.isChecked():
			estado='PENDIENTE'
			fechaini= str(self.fecha_desdelistar.text())
			fechafin=str(self.fecha_hastalistar.text())
			
			tablarecuperada= q.listarporfechaingreso(fechaini,fechafin,concepto,cuenta,estado)
		
			totalfilas= len(tablarecuperada)
			self.tb_listar.setRowCount(totalfilas) #cantidad de filas a llenar *debe ir si o si*
			fila =0
		elif self.rb_observados.isChecked():
			estado='OBSERVADO'
			fechaini= str(self.fecha_desdelistar.text())
			fechafin=str(self.fecha_hastalistar.text())
			
			tablarecuperada= q.listarporfechaingreso(fechaini,fechafin,concepto,cuenta,estado)
		
			totalfilas= len(tablarecuperada)
			self.tb_listar.setRowCount(totalfilas) #cantidad de filas a llenar *debe ir si o si*
			fila =0
			
			
			
		acum=0
		for i in tablarecuperada:
			self.tb_listar.setItem(fila,1,QtGui.QTableWidgetItem(str(i[0])))
			self.tb_listar.setItem(fila,3,QtGui.QTableWidgetItem(str(i[1])))
			self.tb_listar.setItem(fila,6,QtGui.QTableWidgetItem(str(i[2])))
			self.tb_listar.setItem(fila,10,QtGui.QTableWidgetItem(str(i[3])))
			self.tb_listar.setItem(fila,0,QtGui.QTableWidgetItem(str(i[4])))
			self.tb_listar.setItem(fila,7,QtGui.QTableWidgetItem(str(i[5])))
			self.tb_listar.setItem(fila,5,QtGui.QTableWidgetItem(str(i[6])))
			self.tb_listar.setItem(fila,2,QtGui.QTableWidgetItem(str(i[7])))
			self.tb_listar.setItem(fila,8,QtGui.QTableWidgetItem(str(i[8])))
			self.tb_listar.setItem(fila,9,QtGui.QTableWidgetItem(str(i[9])))
			self.tb_listar.setItem(fila,4,QtGui.QTableWidgetItem(str(i[10])))
			self.tb_listar.setItem(fila,11,QtGui.QTableWidgetItem(str(i[13])))
			
			acum=acum+float(i[2])
			
			fila = fila+1
			
	
		
		self.signal_total_listar.setText("${:,}".format(acum).replace(',','~').replace('.',',').replace('~','.'))
		
			
	def exportarlistar(self):
		
		if self.rb_26.isChecked():
			cuenta='26'
		elif self.rb_30.isChecked():
			cuenta='30'
		else:
			cuenta='%'
		
		
		concepto= str(self.cb_concepto_listar.currentText())
		if concepto == '':
			concepto='%'
		
		if self.rb_todos.isChecked():
			estado='%'
			fechaini= str(self.fecha_desdelistar.text())
			fechafin=str(self.fecha_hastalistar.text())
			
			if self.rb_ingresos.isChecked():
				tablarecuperada= q.listarporfechaingreso(fechaini,fechafin,concepto,cuenta,estado)
			else:
				tablarecuperada= q.listarporfechatransferencias(fechaini,fechafin,concepto,cuenta)
		
			totalfilas= len(tablarecuperada)
			self.tb_listar.setRowCount(totalfilas) #cantidad de filas a llenar *debe ir si o si*
			fila =0
		elif self.rb_transferidos.isChecked():
			estado='TRANSFERIDO'
			fechaini= str(self.fecha_desdelistar.text())
			fechafin=str(self.fecha_hastalistar.text())
			
			tablarecuperada= q.listarporfechaingreso(fechaini,fechafin,concepto,cuenta,estado)
		
			totalfilas= len(tablarecuperada)
			self.tb_listar.setRowCount(totalfilas) #cantidad de filas a llenar *debe ir si o si*
			fila =0
		elif self.rb_pendientes.isChecked():
			estado='PENDIENTE'
			fechaini= str(self.fecha_desdelistar.text())
			fechafin=str(self.fecha_hastalistar.text())
			
			tablarecuperada= q.listarporfechaingreso(fechaini,fechafin,concepto,cuenta,estado)
		
			totalfilas= len(tablarecuperada)
			self.tb_listar.setRowCount(totalfilas) #cantidad de filas a llenar *debe ir si o si*
			fila =0
		elif self.rb_observados.isChecked():
			estado='OBSERVADO'
			fechaini= str(self.fecha_desdelistar.text())
			fechafin=str(self.fecha_hastalistar.text())
			
			tablarecuperada= q.listarporfechaingreso(fechaini,fechafin,concepto,cuenta,estado)
		
			totalfilas= len(tablarecuperada)
			self.tb_listar.setRowCount(totalfilas) #cantidad de filas a llenar *debe ir si o si*
			fila =0
		
			
		
		
		
			
			
		
		#EXPORTA A ARCHIVO EXCELL
		
		
		wb = load_workbook('modelo_listar.xlsx')
		#wb = Workbook()
		sheet = wb.active
		
		
		ENCABEZADO = Font(
			name='Calibri',
			size=12,
			bold=True,
			italic=False,
			vertAlign=None,
			underline='none',
			strike=False,
			color='000000')
			
		
		sheet['A1'].font=ENCABEZADO
		sheet['B1'].font=ENCABEZADO
		sheet['C1'].font=ENCABEZADO
		sheet['D1'].font=ENCABEZADO
		sheet['E1'].font=ENCABEZADO
		sheet['F1'].font=ENCABEZADO
		sheet['G1'].font=ENCABEZADO
		sheet['H1'].font=ENCABEZADO
		sheet['I1'].font=ENCABEZADO
		sheet['J1'].font=ENCABEZADO
		sheet['K1'].font=ENCABEZADO
		sheet['L1'].font=ENCABEZADO
		sheet['M1'].font=ENCABEZADO
		sheet['N1'].font=ENCABEZADO
		sheet['O1'].font=ENCABEZADO
		sheet['P1'].font=ENCABEZADO
		
		
		
		
		sheet['A1']="REGISTRO"
		sheet['B1']="DNI"
		sheet['C1']="IMPORTE"
		sheet['D1']="CONCEPTO"
		sheet['E1']="FECHA INGRESO"
		sheet['F1']="FECHA TRANSFERENCIA"
		sheet['G1']="FECHA PRESTACION"
		sheet['H1']="ESTADO"
		sheet['I1']="NUM DE TRANSFERENCIA"
		sheet['J1']="OBS"
		sheet['K1']="NOMBRE"
		sheet['L1']="CUENTA"
		sheet['M1']="INDICE"
		sheet['N1']="ORDEN"
		sheet['O1']="LOTE"
		sheet['P1']="MES"
		
		
		for i in tablarecuperada:
			sheet.append(i)
			
		
		
		wb.save('/home/usuario/listar.xlsx')
		
		msgBox=QtGui.QMessageBox(self.centralwidget)
		msgBox.setIcon(1)
		msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
		msgBox.setWindowTitle("AVISO")
		msgBox.setText("INFOME GENERADO")
		msgBox.exec_()
		
	def limpiarlista(self):
		self.rb_todos.setChecked(True)
		self.rb_t.setChecked(True)
		self.rb_ingresos.setChecked(True)
		self.tb_listar.setRowCount(0)
		self.signal_total_listar.setText("")
		
		
		
	def relaciondegastos(self):
		
		fechaini= str(self.fecha_desdelistar.text())
		fechafin=str(self.fecha_hastalistar.text())
		
		
		if self.rb_26.isChecked():
			cuenta='26'
			tablarecuperada= q.listarpararelacion(fechaini,fechafin,cuenta)
				
			totalbeneficios=0
			totalmp=0
			totaldis=0
			totalafo=0
			totalados=0
			for i in tablarecuperada:
			
				concepto=str(i[2])
				
			
				if concepto =="A. Domiciliario" or concepto=="A. Terapeutico" or concepto=="Geriatria":
					totalados=totalados+float(i[1])
				elif concepto == "Elemento Implantable" or concepto =='Audifono' or concepto=='M. Prestacional':
					totalmp=totalmp+float(i[1])
				elif concepto =="Fonoaudiologia"  or concepto=="Psicoterapia" or concepto=="Transporte" or concepto=="Kinesiologia" or concepto=="T. Ocupacional":
					totaldis=totaldis+float(i[1])
				elif concepto=="Consulta Medica" or concepto=="Analisis" or concepto=="Cobro Indebido" or concepto =="Hoteleria" or concepto =="Traslado":
					totalafo=totalafo+float(i[1])
				elif concepto== "Beneficio de Excepcion":
					totalbeneficios=totalbeneficios+float(i[1])
					
				
			
			
			total= totalbeneficios+totalmp+totaldis+totalafo+totalados
		
			#EXPORTA A ARCHIVO EXCELL
			book = Workbook()
			sheet = book.active
			TITULO = Font(
			name='Calibri',
			size=12,
			bold=True,
			italic=False,
			vertAlign=None,
			underline='none',
			strike=False,
			color='000000')
			
			sheet['A1'].font=TITULO			
		
			sheet['A1']="RELACION DE GASTOS - PRESTACIONAL"
			sheet['A2']="TOTAL POR AF. OBLIGATORIO:"
			sheet['A3']="TOTAL POR DISCAPACIDAD:"
			sheet['A4']="TOTAL POR BENEFICIOS:"
			sheet['A5']="TOTAL POR MEDICO PRESTACIONAL:"
			sheet['A6']="TOTAL SUBSIDIO ADOS"
			sheet['A7']="TOTAL GENERAL"
			sheet['B2']=totalafo
			sheet['B3']=totaldis
			sheet['B4']=totalbeneficios
			sheet['B5']=totalmp
			sheet['B6']=totalados
			sheet['B7']=total
		
				
			sheet['A9']="NOMBRE Y APELLIDO"
			sheet['B9']="IMPORTE"
			sheet['C9']="RUBRO"
			sheet['D9']="FOJA"
		
			cont=0
			subtotal=0
			for i in tablarecuperada:
				sheet.append(i)
				subtotal=subtotal+i[1]
				cont=cont+1
				if cont==50:
					sheet.append(['Subtotal',subtotal])
					sheet.append(['',''])
					sheet.append(['',''])
					sheet.append(['',''])
					sheet.append(['',''])
					sheet.append(['',''])
					sheet.append(['',''])
					sheet.append(['',''])
					sheet.append(['',''])
					sheet.append(['',''])
					sheet.append(['',''])
					sheet.append(['Transporte',subtotal])
					cont=0
			
		
								
			
		
			book.save('/home/usuario/relacion.xlsx')
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(1)
			msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
			msgBox.setWindowTitle("AVISO")
			msgBox.setText("INFOME RELACION DE GASTOS GENERADO")
			msgBox.exec_()
			
					
		elif self.rb_30.isChecked():
			cuenta='30'
			tablarecuperada= q.listarpararelacion(fechaini,fechafin,cuenta)
				
			totalsepelios=0
			totalob=0
			total=0
			for i in tablarecuperada:
			
				concepto=str(i[2])
				
			
				if concepto =="Subsidio de Sepelio":
					totalsepelios=totalsepelios+float(i[1])
				elif concepto == "Ortopedia Blanda":
					totalob=totalob+float(i[1])
				
			
			
			total= totalsepelios+totalob
		
			#EXPORTA A ARCHIVO EXCELL
			book = Workbook()
			sheet = book.active
			TITULO = Font(
			name='Calibri',
			size=12,
			bold=True,
			italic=False,
			vertAlign=None,
			underline='none',
			strike=False,
			color='000000')
			
			sheet['A1'].font=TITULO	
		
			sheet['A1']="RELACION DE GASTOS - SERVICIO COMPLEMENTARIO"
			sheet['A2']="TOTAL POR SEPELIOS:"
			sheet['A3']="TOTAL POR ORTOPEDIA BLANDA:"
			sheet['A4']="TOTAL GENERAL:"
			sheet['B2']=totalsepelios
			sheet['B3']=totalob
			sheet['B4']=total
		
				
			sheet['A7']="NOMBRE Y APELLIDO"
			sheet['B7']="IMPORTE"
			sheet['C7']="RUBRO"
			sheet['D7']="FOJA"
			
			cont=0
			subtotal=0
			for i in tablarecuperada:
				sheet.append(i)
				subtotal=subtotal+i[1]
				cont=cont+1
				if cont==50:
					sheet.append(['Subtotal',subtotal])
					sheet.append(['',''])
					sheet.append(['',''])
					sheet.append(['',''])
					sheet.append(['',''])
					sheet.append(['',''])
					sheet.append(['',''])
					sheet.append(['',''])
					sheet.append(['',''])
					sheet.append(['',''])
					sheet.append(['',''])
					sheet.append(['Transporte',subtotal])
					cont=0
			
					
				
			book.save('/home/usuario/relacion.xlsx')
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(1)
			msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
			msgBox.setWindowTitle("AVISO")
			msgBox.setText("INFOME RELACION DE GASTOS GENERADO")
			msgBox.exec_()	


						
		
			book.save('/home/usuario/relacion.xlsx')
			
			
			
			
			
			
			
			
			
		else:
			cuenta='%'
		
			
			
	def listadoloteables(self):
		fechadesde= str(self.desde_lote.text())
		fechahasta= str(self.hasta_lote.text())
		
		
		
		
		
		if self.rb_sinlote.isChecked():
			listarecuperada=q.listarloteables(fechadesde,fechahasta)
		
		elif self.rb_lotetodos.isChecked(): 
			listarecuperada=q.listarloteablesall(fechadesde,fechahasta)
			
		elif self.rb_xlote.isChecked():
			lote=int(self.txt_filtrado_lotes.text())
			listarecuperada=q.listarloteablesxlote(lote)
			if not listarecuperada:
				msgBox=QtGui.QMessageBox(self.centralwidget)
				msgBox.setIcon(3)
				msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
				msgBox.setWindowTitle("ERROR")
				msgBox.setText("NO EXISTE ESE LOTE")
				msgBox.exec_()
				
		elif self.rb_xafiliado.isChecked():
			afiliado=str(self.txt_filtrado_lotes.text()+'%')
			listarecuperada=q.listarloteablesxafiliado(afiliado)
			
		
		
		totalfilas=len(listarecuperada)
		self.tb_lotes.setRowCount(totalfilas) #cantidad de filas a llenar *debe ir si o si*
		
		fila =0
		acum=0
		cont=0
		
		for i in listarecuperada:
			self.tb_lotes.setItem(fila,0,QtGui.QTableWidgetItem(str(i[0])))
			self.tb_lotes.setItem(fila,1,QtGui.QTableWidgetItem(str(i[1])))
			self.tb_lotes.setItem(fila,2,QtGui.QTableWidgetItem(str(i[2])))
			self.tb_lotes.setItem(fila,3,QtGui.QTableWidgetItem(str(i[3])))
			self.tb_lotes.setItem(fila,4,QtGui.QTableWidgetItem(str(i[4])))
			self.tb_lotes.setItem(fila,5,QtGui.QTableWidgetItem(str(i[5])))
			self.tb_lotes.setItem(fila,6,QtGui.QTableWidgetItem(str(i[6])))
			self.tb_lotes.setItem(fila,7,QtGui.QTableWidgetItem(str(i[7])))
			self.tb_lotes.setItem(fila,8,QtGui.QTableWidgetItem(str(i[8])))
			
			acum=acum+float(i[8])
			cont=cont+1
			fila=fila+1
		self.signal_totallote.setText("${:,}".format(acum).replace(',','~').replace('.',',').replace('~','.'))
		self.signal_cant_lote.setText(str(cont))	
		
	
	def lotear(self):
		indice=int(self.txt_primerindice.text())
		
		indicefin= int(self.txt_ultimoindice.text())
		lote = int(self.txt_numlote.text())
		
		
		lote=q.validalote(lote)
		if lote:
				msgBox=QtGui.QMessageBox(self.centralwidget)
				msgBox.setIcon(3)
				msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
				msgBox.setWindowTitle("ERROR")
				msgBox.setText("ESE NUMERO DE LOTE YA EXISTE")
				msgBox.exec_()
		else:
		
			k = indice
			fallecidos=0
			while k <= indicefin:
				registro=q.traeregistro(k)
				if registro:
					numreg=str(registro[0])
					fecha=str(date.today())
					
					
					mes=str(fecha[5:7])
					anio=str(fecha[0:4])	
												
					estado=q.validaestado(k)
					if estado:			
						if "".join(estado) =='ACTIVO':
							lote = int(self.txt_numlote.text())
							
							q.asignalote(k,lote)
							orden=q.recuperaultimoorden(mes,'26',anio)
							q.asignaloteenregistros(lote,numreg,fecha,int("".join(map(str,orden))),mes,anio)
							
						else:
							#asigna lote 0 a los fallecidos
							q.asignalote(k,'0')
							fallecidos=fallecidos+1
							
					
				k=k+1
		
				
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(1)
			msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
			msgBox.setWindowTitle("LOTE CREADO")
			msgBox.setText("EL LOTE FUE GENERADO CORRECTAMENTE")
			msgBox.exec_()
			
			if fallecidos > 0:
				msgBox=QtGui.QMessageBox(self.centralwidget)
				msgBox.setIcon(1)
				msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
				msgBox.setWindowTitle("REGISTROS OMITIDOS")
				msgBox.setText("SE OMITIERON "+fallecidos+" REGISTROS POR POSEER CUENTAS INACTIVAS")
				msgBox.exec_()
				
		
		#Limpieza
		indice=self.txt_primerindice.setText("")
		indicefin= self.txt_ultimoindice.setText("")
		lote = self.txt_numlote.setText("")
			
	def eliminarloteado(self):
		fila=self.tb_lotes.currentRow()
		item=self.tb_lotes.item(fila,2)
		
		indice=item.text()
		
		msgBox=QtGui.QMessageBox(self.centralwidget)
		msgBox.setIcon(2)
		msgBox.setStandardButtons(QtGui.QMessageBox.Yes | QtGui.QMessageBox.Cancel | QtGui.QMessageBox.No)
		msgBox.setWindowTitle("REGISTRO SELECCIONADO")
		msgBox.setText("QUITAR DE REGISTROS LOTEABLES?")
		r= msgBox.exec_()
		
		
		
		if r==16384:
			
					
			q.eliminarregistroloteable(int(indice))
			
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setWindowTitle(" * * ATENCION * * ")
			msgBox.setText("REGISTRO ELIMINADO")
			resp= msgBox.exec_()
			self.listadoloteables()
			
		elif r==4194304:
			pass
			
		elif r==65536:
			pass
			
		
			
	def eliminar_comision(self):
		fila=self.tb_comisiones.currentRow()
		item1=self.tb_comisiones.item(fila,0)
		item2=self.tb_comisiones.item(fila,1)
		
		fecha=item1.text()
		agente=item2.text()
		
		msgBox=QtGui.QMessageBox(self.centralwidget)
		msgBox.setIcon(2)
		msgBox.setStandardButtons(QtGui.QMessageBox.Yes | QtGui.QMessageBox.Cancel | QtGui.QMessageBox.No)
		msgBox.setWindowTitle("COMISION SELECCIONADA")
		msgBox.setText("ELIMINAR COMISION?")
		r= msgBox.exec_()
		
		
		
		if r==16384:
			
					
			q.eliminarcomision(str(fecha),str(agente))
			
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setWindowTitle(" * * ATENCION * * ")
			msgBox.setText("REGISTRO ELIMINADO")
			resp= msgBox.exec_()
			self.vercomisiones()
			
		elif r==4194304:
			pass
			
		elif r==65536:
			pass		
			
	def exportarlote(self):
		fechadesde= str(self.desde_lote.text())
		fechahasta= str(self.hasta_lote.text())
		
		
		listarecuperada=q.listarloteablesforexport(fechadesde,fechahasta)
		
		book = Workbook()
		sheet = book.active
		
		sheet['A1']="SISTEMA"
		sheet['B1']="SUCURSAL"
		sheet['C1']="CUENTA"
		sheet['D1']="LOTE"
		sheet['E1']="IMPORTE"
		
		
		for i in listarecuperada:
			sheet.append(i)
		
				
		book.save('/home/usuario/lote.xlsx')
			
	def traeultimolote(self):
		ultimolote=q.recuperaultimolote()
		self.signal_ultlote.setText(str(ultimolote[0]))	
	
	def exportardetallelote(self):
		
		fechadesde= str(self.desde_lote.text())
		fechahasta= str(self.hasta_lote.text())
		
		
		listarecuperada=q.listardetalleforexport(fechadesde,fechahasta)
		total=str(self.signal_totallote.text())
		book = Workbook()
		sheet = book.active
		sheet['A1']="TOTAL LOTE"
		sheet['B1']=total
		
		sheet['A2']="REGISTRO"
		sheet['B2']="DNI"
		sheet['C2']="IMPORTE"
		sheet['D2']="NOMBRE"
		
		
		
		for i in listarecuperada:
			sheet.append(i)
		
				
		book.save('/home/usuario/detallelote.xlsx')
		self.traeultimolote()
		
		msgBox=QtGui.QMessageBox(self.centralwidget)
		msgBox.setIcon(1)
		msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
		msgBox.setWindowTitle("INFORMACION")
		msgBox.setText("DETALLE EXPORTADO")
		msgBox.exec_()
	
	
		
		
	def altaafiliado(self):
		
		
		if self.txt_dnialta.text() or self.txt_nombrealta.text():
		
			dni= int(self.txt_dnialta.text())
			nombre= str(self.txt_nombrealta.text())
			email=str(self.txt_emailalta.text())
			cbu=str(self.txt_cbualta.text())
			estado='ACTIVO'
									
			
			cuenta=int(cbu[12:21]) # dentro del cbu esta incluido el numero de cuenta
			sucursal=int(cbu[9:12]) #dentro del cbu tambien esta incluida la sucursal
			
			a=cbu[0:2] #CHEQUEO SI LOS PRIMEROS 2 NUMEROS DEL CBU SON 33 para determinar si es sistema 1 o 0
			
			if a == '33':
				sistema= 1
			else:
				sistema= 0
			
			blanco=0
			
			for i in cbu:
				
				if i==' ':
					blanco=blanco+1
					
				
					
			
			tienecuenta=q.validarcuentas(dni)
			cbuexiste=q.validarcbu(cbu)
			
			if blanco==0 and len(cbu)==22:
			
				if not tienecuenta:
					if not cbuexiste:
					
				
						q.ingresarafiliado(dni,nombre,cbu,cuenta,email,sistema,sucursal,estado)
						msgBox=QtGui.QMessageBox(self.centralwidget)
						msgBox.setIcon(1)
						msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
						msgBox.setWindowTitle("EXITO")
						msgBox.setText("ALTA OK!")
						msgBox.exec_()
						self.txt_dnialta.setFocus()
					else:
						msgBox=QtGui.QMessageBox(self.centralwidget)
						msgBox.setIcon(3)
						msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
						msgBox.setWindowTitle("ERROR GRAVE")
						msgBox.setText("CBU CORRESPONDE A OTRO AFILIADO")
						msgBox.exec_()
						self.txt_dnialta.setFocus()
						
					
				else:
					
					msgBox=QtGui.QMessageBox(self.centralwidget)
					msgBox.setIcon(3)
					msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
					msgBox.setWindowTitle("AFILIADO EXISTENTE")
					msgBox.setText("YA SE HA INGRESADO ESE AFILIADO CON ANTERIORIDAD")
					msgBox.exec_()
					self.txt_dnialta.setFocus()
			
			else:
				msgBox=QtGui.QMessageBox(self.centralwidget)
				msgBox.setIcon(3)
				msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
				msgBox.setWindowTitle("MAL CBU")
				msgBox.setText("FORMATO CBU INCORRECTO")
				msgBox.exec_()
				
				
			
		else:
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(3)
			msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
			msgBox.setWindowTitle("ERROR")
			msgBox.setText("CAMPOS VACIOS")
			msgBox.exec_()
			self.txt_dnialta.setFocus()
		
			
			
		
	def buscarafiliado(self):
		
		campo= int(self.txt_busqafiliado.text())
		
		
		
		
		if self.rb_dni_busafi.isChecked():
			if campo=='':
				campo=0
			afiliado = q.busquedaafiliado(campo)
		else:
			if campo=='':
				campo=0
			afiliado =q.busquedaafiliadoxcta(campo)
			
			
		
		
		if afiliado:
		
			#TARJETA
			self.signal_nombre.setText(str(afiliado[1]))
			self.signal_cuil.setText(str(afiliado[8]))
			self.signal_cbu.setText(str(afiliado[2]))
			self.signal_email.setText(str(afiliado[7]))
		
			#CAMPOS
			self.txt_nombre_busafi.setText(str(afiliado[1]))
			self.txt_cbu_busafi.setText(str(afiliado[2]))
			self.txt_telbusafi.setText(str(afiliado[10]))
			self.txt_email_busafi.setText(str(afiliado[7]))
			self.txt_aliasbusafi.setText(str(afiliado[9]))
			self.txt_cuil_busafi.setText(str(afiliado[8]))
			self.txt_cuenta_busafi.setText(str(afiliado[3]))
			self.combo_estado_busafi.setText(str(afiliado[6]))
			estado =str(afiliado[6])
			
			print estado
			if estado =='FALLECIDO':
				self.cbx_inactivo.setChecked(True)
				
		else:
			#TARJETA
			self.signal_nombre.setText("NO ENCONTRADO")
			self.signal_cuil.setText("NO ENCONTRADO")
			self.signal_cbu.setText("NO ENCONTRADO")
			self.signal_email.setText("NO ENCONTRADO")
			
			
		
	def updateafiliado(self):
		
				
		
		#GUARDAR DATOS EN VARIABLES
		dni = int(self.txt_busqafiliado.text())
		nombre = str(self.txt_nombre_busafi.text())
		cbu= str (self.txt_cbu_busafi.text())
		cuenta= int (self.txt_cuenta_busafi.text())
		estado = str(self.combo_estado_busafi.text())
		email = str(self.txt_email_busafi.text())
		#alias = str(self.txt_aliasbusafi.text())
		#tel= self.txt_telbusafi.text()
		#cuil= self.txt_cuil_busafi.text()
		
			
		
		
		if dni =='':
			dni=0
		if nombre =='':
			nombre=''
		if cbu=='':
			cbu=0
		if cuenta=='':
			cuenta=0
		if email=='':
			email=''
		'''if alias=='':
			alias=''
		if tel =='':
			tel=0
		if cuil =='':
			cuil=0'''
			
		if self.cbx_inactivo.isChecked():
			estado='FALLECIDO'
				
		
		
		if self.rb_dni_busafi.isChecked():
			
			q.actualizaafiliado(nombre,cbu,cuenta,estado,email,dni)
			
		else:
			q.actualizaafiliadoxcta(nombre,cbu,cuenta,estado,email,dni)
			
	
	def eliminarafiliados(self):
		
		msgBox=QtGui.QMessageBox(self.centralwidget)
		msgBox.setIcon(2)
		msgBox.setStandardButtons(QtGui.QMessageBox.Yes | QtGui.QMessageBox.Cancel | QtGui.QMessageBox.No)
		msgBox.setWindowTitle("ELIMINAR AFILIADO")
		msgBox.setText("DESEA ELIMINAR EL AFILIADO?")
		r= msgBox.exec_()
		
		
		
		if r==16384:
			
			
		
			dni=int(self.txt_busqafiliado.text())
		
			q.eliminarafiliado(dni)
			
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setWindowTitle(" * * ATENCION * * ")
			msgBox.setText("REGISTRO ELIMINADO")
			resp= msgBox.exec_()
			
		elif r==4194304:
			pass
			
		elif r==65536:
			pass
		
	
		
		
		
		
	def nuevacomision(self):
		
			
		destino=str(self.txt_destino.currentText())
		localidad=str(self.txt_localidad_comision.currentText())
		fecha_comision= str(self.date_comision.text())
		agente=str(self.cb_agente.currentText())
		valida_comision=q.validacomision(fecha_comision,agente)
		
		if valida_comision=='0':
			
			if self.cbx_eneldia.isChecked():
				fecha_vuelta=fecha_comision
			else:
				fecha_vuelta=str(self.date_comision_regreso.text())
				
			fecha_hoy=str(date.today())
			year=int(fecha_comision[6:10])
			month=int(fecha_comision[3:5])
			
			if destino =="BS. AS." or destino == "Entre Rios":
				valor_comision = float("".join(map(str,q.traervalorcomisionaf(year,month))))
			else:
				
				valor_comision = float("".join(map(str,q.traervalorcomision(year,month))))
				
			
			
			
					
			hora_fin=str(self.txt_fin_comi.text())
			hora_inicio=str(self.txt_ini_comi.text())
			destino=str(self.txt_destino.currentText())
			transporte=str(self.cbx_transporte.currentText())
			diferencia =int(fecha_vuelta[0:2])-int(fecha_comision[0:2])+1
			
			
			
			
			
			if str(self.cbx_transporte.currentText())=="Oficial":
				valor_pasaje=0
				articulo=0
			else:
				valor_pasaje=float(self.txt_valor.text())
				articulo=float(valor_comision*0.10) #articulo 6 ley 79.14
				
			razon=str(self.txt_razon.currentText())
				
			importe =valor_comision+valor_pasaje+articulo
			
			q.insertarcomision(fecha_comision,hora_inicio,hora_fin,destino,transporte,razon,importe,agente,fecha_vuelta,localidad,valor_pasaje)
			self.vercomisiones()
			
			
			#EXPORTAR A EXCEL
			
			wb = load_workbook('viatico.xlsx')
			
			sheet = wb.active
			
			
			sheet['B7']=agente
			sheet['B9']=localidad
			sheet['B11']=razon
			sheet['B13']=transporte
			#sheet['B27']=fecha_hoy
			
			
			sheet['A19']=str(fecha_comision)
			sheet['C19']=str(hora_inicio)
			sheet['D19']=str(fecha_vuelta)
			sheet['F19']=str(hora_fin)
			sheet['G19']=str(diferencia)
			sheet['H19']=str(valor_comision)
			sheet['J19']=str("Comision")
			sheet['H20']=str(valor_pasaje)
			sheet['J20']=str("Pasaje")
			sheet['H21']=str(articulo)
			sheet['J21']=str("Art. 6 Ley 79.14")
			sheet['I22']=str(importe)
			
		
			path ="/home/usuario/comisiones/comision_generada {}.xlsx".format(agente + fecha_comision)
			
			wb.save(path)
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(1)
			msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
			msgBox.setWindowTitle("AVISO")
			msgBox.setText("COMISION LIQUIDADA")
			msgBox.exec_()
		
		else:
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(3)
			msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
			msgBox.setWindowTitle("COMISION EXISTENTE")
			msgBox.setText("YA EXISTE UNA COMISION PARA ESA FECHA Y PARA ESE AGENTE")
			msgBox.exec_()
			self.txt_dnialta.setFocus()
			
			
	def ajustar_comision(self):
		fila=self.tb_comisiones.currentRow()
		item1=self.tb_comisiones.item(fila,0)
		item2=self.tb_comisiones.item(fila,1)
		item3=self.tb_comisiones.item(fila,3)
		item4=self.tb_comisiones.item(fila,4)
		item8=self.tb_comisiones.item(fila,8)
		
		fecha=item1.text()
		agente=item2.text()
		importe=float(item3.text())
		transporte=item4.text()
		pasaje=float(item8.text())
		
		valor_ajuste=float(self.txt_valor_ajuste.text())
		
		
		if transporte == 'Oficial':
			diferencia=float(valor_ajuste)-importe
		else:
			diferencia=valor_ajuste-importe + (valor_ajuste-importe*0.10)-pasaje
			
			
		#EXPORTAR A EXCEL
			
		wb = load_workbook('viatico.xlsx')
			
		sheet = wb.active
			
			
		sheet['B7']=str(agente)
			
		sheet['B13']=str(transporte)
			
			
			
		sheet['A19']=str(fecha)
			
			
			
		sheet['H19']=str(diferencia)
		sheet['J19']=str("Dif. por ajuste")
			
			
			
		
		path ="/home/usuario/comisiones/comision_generada {}.xlsx".format("AJUSTE "+agente + fecha)
			
		wb.save(path)
		msgBox=QtGui.QMessageBox(self.centralwidget)
		msgBox.setIcon(1)
		msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
		msgBox.setWindowTitle("AVISO")
		msgBox.setText("COMISION AJUSTADA")
		msgBox.exec_()
		
		
		
		
		
				
		
	def vercomisiones(self):
		
		nombre=str(self.cb_agente.currentText())
		comisiones=q.traercomisiones(nombre)
		
		
			
		totalfilas=len(comisiones)
		self.tb_comisiones.setRowCount(totalfilas) #cantidad de filas a llenar *debe ir si o si*
		
		fila =0
		acum=0
		
		for i in comisiones:
			self.tb_comisiones.setItem(fila,0,QtGui.QTableWidgetItem(str(i[0])))
			self.tb_comisiones.setItem(fila,1,QtGui.QTableWidgetItem(str(i[1])))
			self.tb_comisiones.setItem(fila,2,QtGui.QTableWidgetItem(str(i[2])))
			self.tb_comisiones.setItem(fila,3,QtGui.QTableWidgetItem(str(i[3])))
			self.tb_comisiones.setItem(fila,4,QtGui.QTableWidgetItem(str(i[4])))
			self.tb_comisiones.setItem(fila,5,QtGui.QTableWidgetItem(str(i[5])))
			self.tb_comisiones.setItem(fila,6,QtGui.QTableWidgetItem(str(i[6])))
			self.tb_comisiones.setItem(fila,7,QtGui.QTableWidgetItem(str(i[7])))
			self.tb_comisiones.setItem(fila,8,QtGui.QTableWidgetItem(str(i[8])))
			
			
			
			fila=fila+1
			
	
			
	def nuevoagente(self):
		nombre = str(self.txt_agregar_agente.text())
		dni = int(self.txt_agregar_dniagente.text())
		
		existe=q.validaragente(dni)
		if existe:
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(3)
			msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
			msgBox.setWindowTitle("ERROR")
			msgBox.setText("ESE AGENTE YA FUE INGRESADO CON ANTERIORIDAD")
			msgBox.exec_()
			
		else:
			q.insertaragente(nombre,dni)
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(1)
			msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
			msgBox.setWindowTitle("OK")
			msgBox.setText("ALTA REALIZADA")
			msgBox.exec_()
		
		self.txt_agregar_agente.setText("")
		self.txt_agregar_dniagente.setText("")
		self.cb_agente.clear()
		self.cb_agente.addItem("")
		self.llenarcombos()
		
		
	def limpiar_registropago(self):
		self.txt_acttransferencia.setText("")
		self.txt_actregistro.setText("")
		self.txt_busqueda.setText("")
		self.txt_actregistro.setText("")
		self.txt_actindice.setText("")
		self.txt_actorden.setText("")			
		
			
					
	def limpiar(self):
		self.txt_registro.setText("")
		self.txt_dni.setText("")
		self.txt_observaciones.setText("")
		self.txt_importe.setText("")
		self.txt_ayn.setText("")
		#self.cb_ingresos.clean()		
		
		
		
		#PAGINA BUSCAR
		self.txt_busqueda.setText("")
		#self.tb_busqueda.deleteAllRows()
		
		#pagina actualizar
		self.txt_modificar_registro.setText("")
		self.txt_modificar_dni.setText("")
		self.txt_modificar_nombre.setText("")
		self.txt_modificar_importe.setText("")
		self.txt_modificar_concepto.setText("")
		self.txt_modificar_fechapago.setText("")
		self.txt_modificar_fechaprestacion.setText("")
		self.txt_modificar_transferencia.setText("")
		self.txt_modificar_obs.setText("")
		
		#bloquea campos actualizar
		self.txt_modificar_dni.setEnabled(False) 
		self.txt_modificar_nombre.setEnabled(False)
		self.txt_modificar_importe.setEnabled(False)
		self.txt_modificar_concepto.setEnabled(False)
		self.txt_modificar_fechapago.setEnabled(False)
		self.txt_modificar_fechaprestacion.setEnabled(False)
		self.txt_modificar_transferencia.setEnabled(False)
		self.txt_modificar_obs.setEnabled(False)
		self.txt_modificar_cuenta.setEnabled(False)
		self.txt_modificar_nuevoreg.setEnabled(False)
		
		#bloque campos buscar afiliado
		self.txt_nombre_busafi.setEnabled(False)
		self.txt_cbu_busafi.setEnabled(False)
		self.txt_telbusafi.setEnabled(False)
		self.txt_email_busafi.setEnabled(False)
		self.txt_aliasbusafi.setEnabled(False)
		self.txt_cuenta_busafi.setEnabled(False)
		self.txt_cuil_busafi.setEnabled(False)
		self.combo_estado_busafi.setEnabled(False)
		
		
		
		
		
		self.txt_dnialta.setText("")
		self.txt_nombrealta.setText("")
		
		self.txt_emailalta.setText("")
		
		
		self.txt_cbualta.setText("")
		
		
	def bloquearrblistar(self):
		self.rb_transferidos.hide()
		self.rb_pendientes.hide()
		self.rb_observados.hide()
	
	def mostrarrblistar(self):
		self.rb_transferidos.show()
		self.rb_pendientes.show()
		self.rb_observados.show()
		
		
	def importar(self):
		
		ruta=str(self.txt_importar_registros.text())
		wb = load_workbook(ruta)
		
		#contar filas de excel
		sheet = wb.active
		totalfilas = sheet.max_row
		
		nocuenta=0
		duplicados=0
		registrados=0
		for i in range(1,totalfilas+1):
			
						
			fecha=str(sheet['A{}'.format(i)].value)
			registro=str(sheet['C{}'.format(i)].value)
			dni=int(sheet['E{}'.format(i)].value)
			importe =float(sheet['K{}'.format(i)].value)
			ayn=str(sheet['F{}'.format(i)].value)
			year=str(sheet['M{}'.format(i)].value)
			month=str(sheet['L{}'.format(i)].value)
			fechap=str(month+year)
			estado=str(sheet['D{}'.format(i)].value)
			
			observaciones=str(sheet['T{}'.format(i)].value)
			concepto= sheet['N{}'.format(i)].value
			
			if concepto == "SSE":
				concepto="Subsidio de Sepelio"
				cuenta=30
			elif concepto == "BE":
				concepto="Beneficio de Excepcion"
				cuenta=26
			elif concepto =="DIS":
				concepto ="Beneficio de Excepcion"
				cuenta=26
			elif concepto =="AFO":
				concepto="Cobro Indebido"
				cuenta=26
			elif concepto =="AT":
				concepto = "A. Terapeutico"
				cuenta=26
			elif concepto =="EM":
				concepto = "Ortopedia Blanda"
				cuenta=30
			elif concepto =="MP":
				concepto = "Elemento Implantable"
				cuenta=26
			
			
			if self.cbx_imp_registros.isChecked(): #LOTEABLES CHECKED
				tienecuenta=q.validarcuentas(dni)
				if tienecuenta != None:
					q.insertarenbd(registro,dni,importe,concepto,fecha,fechap,estado,observaciones,ayn,cuenta)
					q.insertarenlotes(registro,dni,importe,fecha)
					registrados=registrados+1
										
					
				else:
					print "no se puede lotear el afiliado "+str(ayn)+", no tiene cuenta"
					nocuenta=nocuenta+1
					
			else:
				duplicado=q.validarduplicados(registro)
				if duplicado !=None:
					print "ya existe ese registro"
					duplicados=duplicados+1
				else:
					q.insertarenbd(registro,dni,importe,concepto,fecha,fechap,estado,observaciones,ayn,cuenta)
					registrados=registrados+1
						
		msgBox=QtGui.QMessageBox(self.centralwidget)
		msgBox.setIcon(1)
		msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
		msgBox.setWindowTitle("REGISTRO IMPORTADOS")
		msgBox.setText("SE HAN IMPORTADO "+""+str(registrados)+" REGISTROS"+
		"NO SE HAN IMPORTADO"+""+str(duplicados)+" REGISTROS (DUPLICADOS)"
		+"y"+str(nocuenta)+" debido a que no tenian cuenta")
		msgBox.exec_()
	
	def importarcuentas(self):
		
		ruta=str(self.txt_archivo_cuentas.text())
		wb = load_workbook(ruta)
		#contar filas excell para determinar cantidad a recorrer por el bucle for
		
		sheet = wb.active
		totalfilas = sheet.max_row
		repetidos=0
		validas=0
		for i in range(1,totalfilas+1):
			
						
			dni=int(sheet['E{}'.format(i)].value)
			nombre=str(sheet['F{}'.format(i)].value)
			sistema=int(sheet['G{}'.format(i)].value)
			sucursal=int(sheet['H{}'.format(i)].value)
			cuenta=int(sheet['I{}'.format(i)].value)
			cbu=0
			email=0
			telefono=0
			alias=0
			cuil=0
			
			
			
			#VALIDO QUE NO EXISTA LA CUENTA
			cuentas=q.validarcuentas(dni)
			
			if cuentas == None:
			
				q.ingresarafiliado(dni,nombre,cbu,cuenta,alias,email,telefono,cuil,sistema,sucursal)
				validas=validas+1
			else:
				print "cuenta existente"+" "+str(dni)+" "+nombre
				repetidos=repetidos+1
		
		msgBox=QtGui.QMessageBox(self.centralwidget)
		msgBox.setIcon(1)
		msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
		msgBox.setWindowTitle("REGISTRO IMPORTADOS")
		msgBox.setText("SE HAN IMPORTADO "+""+str(validas)+" REGISTROS"+
		" NO SE HAN IMPORTADO"+""+str(repetidos)+" REGISTROS (DUPLICADOS)")
		
		msgBox.exec_()
		
	def importarpagos(self):
		noencontrados=[]
		nocoinciden=[]
		
		ruta=str(self.txt_archivo_pagos.text())
		wb = load_workbook(ruta)
		#contar filas excell para determinar cantidad a recorrer por el bucle for
		sheet = wb.active
		totalfilas = sheet.max_row
		
		cont_actualizados=0
		cont_noactualizados=0
		cont_inconsistentes=0
			
		if self.cbx_desdebsf.isChecked():
		#DESDE BANCO
			for i in range(7,totalfilas+1):
				registro=str(sheet['K{}'.format(i)].value)
				existe=q.validaregistro(registro)
				if existe != None:
					importe=float(sheet['j{}'.format(i)].value)
					fechapago=str(sheet['D{}'.format(i)].value)
					orden=int(sheet['R{}'.format(i)].value)			
					mes=str(fechapago[5:7])
					anio=str(fechapago[0:4])
					transferencia=int(sheet['A{}'.format(i)].value)
					estado="TRANSFERIDO"	
					fecha=fechapago[0:10]
					q.actualizapagoenbd_masivo(registro,fecha,transferencia,estado,orden,mes,anio)
					cont_actualizados = cont_actualizados+1
					
					#comparar montos de excel contra registro
					if existe[1] != importe:
						
						nocoinciden.append(registro)
						cont_inconsistentes +=1
					
				else:
					
					noencontrados.append(registro)
					cont_noactualizados= cont_noactualizados=+1
					
		else:
			for i in range(1,totalfilas+1):
				registro=str(sheet['C{}'.format(i)].value)
				dni=int(sheet['E{}'.format(i)].value)
				print dni
				existe=q.validaregistro(registro,dni)
				if existe != None:
					importe=float(sheet['K{}'.format(i)].value)
					fecha=str(sheet['P{}'.format(i)].value)
					orden=int(sheet['S{}'.format(i)].value)			
					mes=str(fecha[5:7])
					transferencia=int(sheet['Q{}'.format(i)].value)
					estado="TRANSFERIDO"	
					
					q.actualizapagoenbd_masivo(registro,fecha,transferencia,estado,orden,mes,dni)
					print "actualizado"
					cont_actualizados = cont_actualizados+1
				else:
					print "registro no encontrado "+str(registro)
					cont_noactualizados= cont_noactualizados=+1
					noencontrados.append(registro)
		
		if cont_noactualizados >0 or cont_inconsistentes >1:
			
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(1)
			msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
			msgBox.setWindowTitle("PAGOS ACTUALIZADOS")
			msgBox.setText("VER REPORTE!")
			
			msgBox.exec_()
		else:
			
			
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(1)
			msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
			msgBox.setWindowTitle("PAGOS ACTUALIZADOS")
			msgBox.setText("OK! TODOS LOS REGISTROS ACTUALIZADOS")
			
			msgBox.exec_()
		
		book = Workbook()
		sheet = book.active
		
		
		ENCABEZADO = Font(
			name='Calibri',
			size=12,
			bold=True,
			italic=False,
			vertAlign=None,
			underline='none',
			strike=False,
			color='000000')
			
		
		sheet['A1'].font=ENCABEZADO
		
		
		sheet['A1']="REPORTE DE REGISTROS IMPORTADOS"
		sheet['A2']="Registros Actualizados"
		sheet['B2']=str(cont_actualizados)
		
		sheet['A3']="Registros NO Actualizados"
		sheet['B3']=str(cont_noactualizados)
		
		sheet['A4']="Registros Inconsistentes"
		sheet['B4']=str(cont_inconsistentes)
		
		
		sheet['A5']="NO ENCONTRADOS:"
		sheet.append(noencontrados)
		sheet['A7']="INCONSISTENTES:"
		
		sheet.append(nocoinciden)
		
		book.save('/home/usuario/no_encontrados.xlsx')
		
					
		#DESDE PLANILLA EXCEL
	
	def avisomail(self):
		
		registro=self.txt_actregistro.text()
		dni_afiliado=q.recuperaDniAfiliado(str(registro))
		email_afiliado=q.recuperaEmailAfiliado(dni_afiliado)
		datos_registro=q.recuperadatosenbd(str(registro))
		print "".join(map(str,dni_afiliado))
		print email_afiliado
		if email_afiliado:
			print "".join(map(str,email_afiliado))
		else:
			email_afiliado='zenfranco@gmail.com'
			print "".join(map(str,email_afiliado))
			
		
		for i in datos_registro:
			print i[1] #nombre
			print i[2] #importe
			print i[3] #rubro
			nombre=i[1]
			importe=i[2]
			rubro=i[3]
		
		# Create an email message object 
		message = MIMEMultipart()
		
		
		
		email_subject = "AVISO DE TRANSFERENCIA - IAPOS | Habilitacion"
		sender_email_address = "iapos_ashabiltacion@santafe.gov.ar" 
		receiver_email_address = email_afiliado

		# Configure email headers 
		message['Subject'] = email_subject 
		message['From'] = sender_email_address 
		message['To'] = receiver_email_address
		
		#message.set_content("Habilitacion informa pago de expediente por el valor de"+importe)
		email_smtp = "correo.santafe.gov.ar"  
		
				
		# Set smtp server and port 
		server = smtplib.SMTP(email_smtp, '587') 

		# Identify this client to the SMTP server 
		server.ehlo() 

		# Secure the SMTP connection 
		server.starttls()
		
		sender_email_address = "iapos_ashabilitacion@santafe.gov.ar" 
		email_password = "habi2017" 

		# Login to email account 
		server.login(sender_email_address, email_password) 

		# Send email 
		#server.send_message(message)
		

		# Close connection to server 
		#server.quit()
		s = smtplib.SMTP('correo.santafe.gov.ar')
		s.sendmail("iapos_ashabiltacion@santafe.gov.ar", [email_afiliado], message.as_string()+"IAPOS INFORMA")
		s.quit()
						
			
		
	def definirvalorcomision(self):
		fecha_comision=self.fecha_valor_comision.text()
		valor=float(self.txt_valor_comision.text())
		fecha_vigencia="01-"+str(fecha_comision)
		year=str(fecha_comision[3:7])
		month=str(fecha_comision[0:2])
		
		valida=q.validarvalores(month,year)
		if valida:
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(2)
			msgBox.setStandardButtons(QtGui.QMessageBox.Yes | QtGui.QMessageBox.Cancel | QtGui.QMessageBox.No)
			msgBox.setWindowTitle("AVISO")
			msgBox.setText("YA EXISTE UN VALOR DEFINIDO PARA ESE MES. DESEA MODIFICAR EL VALOR?")
			r= msgBox.exec_()
			
			
			
			if r==16384:
				
				
				
				q.updatevalorcomision(valor,int(month),int(year))
				
				msgBox=QtGui.QMessageBox(self.centralwidget)
				msgBox.setWindowTitle(" * * ATENCION * * ")
				msgBox.setText("VALOR MODIFICADO")
				resp= msgBox.exec_()
				
			elif r==4194304:
				pass
				
			elif r==65536:
				pass
		else:
			q.nuevovalorcomision(int(year),int(month),valor,fecha_vigencia)
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(1)
			msgBox.setWindowTitle("AVISO ")
			msgBox.setText("VALOR DEFINIDO CORRECTAMENTE")
			resp= msgBox.exec_()
		
		
		
		
				
#SELECTOR DE ARCHIVOS
	def elegirarchivo(self):
		self.txt_archivo_cuentas.setText(QtGui.QFileDialog.getOpenFileName())
		
	def elegirarchivo_registros(self):
		self.txt_importar_registros.setText(QtGui.QFileDialog.getOpenFileName())
	
	def elegirarchivo_pagos(self):
		self.txt_archivo_pagos.setText(QtGui.QFileDialog.getOpenFileName())

				
			
	
	
	def fsalir(self):
		self.close()
	
	
	






#INICIO DE PROGRAMA

if __name__ == '__main__':
	q=base()
	app = QtGui.QApplication(sys.argv)
	MyWindow = VentanaPrincipal(None)
	MyWindow.show()
	MyWindow.llenarcombos()
	
	app.exec_()







